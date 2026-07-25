"""openpyxl-based export of the official annual report layout: merged
bold section headers, numbered lines, indented subset lines, "Jami" rows,
and the Scopus/WoS quartile breakdown for line 2.1."""

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

SECTION_FILL = PatternFill(start_color="0F2A43", end_color="0F2A43", fill_type="solid")
TOTAL_FILL = PatternFill(start_color="F6EDE0", end_color="F6EDE0", fill_type="solid")
SECTION_FONT = Font(bold=True, color="FFFFFF")
TOTAL_FONT = Font(bold=True)
HEADER_FONT = Font(bold=True, size=13)


def _write_header_block(ws, row, institute_name, period_label, employee_info=None):
    ws.cell(row=row, column=1, value=institute_name).font = HEADER_FONT
    row += 1
    ws.cell(row=row, column=1, value=f"Hisobot davri: {period_label}")
    row += 1
    if employee_info:
        ws.cell(row=row, column=1, value=f"F.I.Sh.: {employee_info['full_name']}")
        row += 1
        ws.cell(row=row, column=1, value=(
            f"Ilmiy daraja: {employee_info.get('academic_degree', '')} · "
            f"Lavozim: {employee_info.get('position', '')} · "
            f"Laboratoriya: {employee_info.get('department', '')}"
        ))
        row += 1
    return row + 1


def _write_report_sections(ws, row, report):
    col_line, col_label, col_count = 1, 2, 3

    for section in report["sections"]:
        ws.cell(row=row, column=1, value=f"{section['id']}.  {section['title']}")
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
        for c in range(1, 4):
            ws.cell(row=row, column=c).fill = SECTION_FILL
            ws.cell(row=row, column=c).font = SECTION_FONT
        row += 1

        for line in section["lines"]:
            indent = "    " if line.get("is_subset") else ""
            ws.cell(row=row, column=col_line, value=line["code"])
            label_cell = ws.cell(row=row, column=col_label, value=f"{indent}{line['label']}")
            if line.get("is_subset"):
                label_cell.font = Font(italic=True, color="4B5C72")
            ws.cell(row=row, column=col_count, value=line["count"])

            # Quartile breakdown directly under line 2.1.
            if line["code"] == "2.1":
                row += 1
                qm = report["quartile_matrix"]
                ws.cell(row=row, column=col_label, value="        Scopus: Q1/Q2/Q3/Q4")
                for i, q in enumerate(["Q1", "Q2", "Q3", "Q4"]):
                    ws.cell(row=row, column=col_count + 1 + i, value=qm["scopus"][q])
                row += 1
                ws.cell(row=row, column=col_label, value="        Web of Science: Q1/Q2/Q3/Q4")
                for i, q in enumerate(["Q1", "Q2", "Q3", "Q4"]):
                    ws.cell(row=row, column=col_count + 1 + i, value=qm["wos"][q])
            row += 1

        ws.cell(row=row, column=col_label, value="Jami:")
        ws.cell(row=row, column=col_count, value=section["total"])
        for c in range(1, 4):
            ws.cell(row=row, column=c).fill = TOTAL_FILL
            ws.cell(row=row, column=c).font = TOTAL_FONT
        row += 2

    if report["without_certificate"]:
        ws.cell(row=row, column=1, value=f"* {report['without_certificate']} ta yozuv sertifikatsiz (IV bo'lim).")
        row += 1
    if report["unclassified"]:
        ws.cell(row=row, column=1, value=(
            f"* {len(report['unclassified'])} ta yozuv hisobotga kiritilmadi "
            f"(klassifikatsiya maydonlari to'ldirilmagan)."
        ))
        row += 1

    return row


def build_personal_report_workbook(report, period_label, employee_info):
    wb = Workbook()
    ws = wb.active
    ws.title = "Hisobot"
    for i, width in enumerate([10, 55, 12, 8, 8, 8, 8], start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    row = _write_header_block(ws, 1, "Seysmologiya instituti", period_label, employee_info)
    _write_report_sections(ws, row, report)
    return wb


def build_institute_report_workbook(report, period_label, department_label=None):
    wb = Workbook()
    ws = wb.active
    ws.title = "Institut hisoboti"
    for i, width in enumerate([10, 55, 12, 8, 8, 8, 8], start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    row = _write_header_block(ws, 1, f"Seysmologiya instituti — {department_label or 'butun institut'}", period_label)
    if report.get("deduplicated_count"):
        ws.cell(row=row, column=1, value=f"{report['deduplicated_count']} ta takroriy yozuv birlashtirildi.")
        row += 2
    row = _write_report_sections(ws, row, report)

    # Per-laboratory breakdown sheet.
    if report.get("department_breakdown"):
        ws2 = wb.create_sheet("Laboratoriyalar bo'yicha")
        ws2.cell(row=1, column=1, value="Laboratoriya").font = TOTAL_FONT
        codes = ["total_II", "total_III", "total_IV", "total_V", "total_VI"]
        for i, code in enumerate(codes, start=2):
            ws2.cell(row=1, column=i, value=code.replace("total_", "Bo'lim "))
        for r, entry in enumerate(report["department_breakdown"], start=2):
            ws2.cell(row=r, column=1, value=entry["department"])
            for i, code in enumerate(codes, start=2):
                ws2.cell(row=r, column=i, value=entry["report"].get(code, 0))
        ws2.column_dimensions["A"].width = 30

    # Per-employee breakdown sheet.
    if report.get("employee_breakdown"):
        ws3 = wb.create_sheet("Xodimlar bo'yicha")
        ws3.cell(row=1, column=1, value="F.I.Sh.").font = TOTAL_FONT
        ws3.cell(row=1, column=2, value="Laboratoriya").font = TOTAL_FONT
        codes = ["total_II", "total_III", "total_IV", "total_V", "total_VI"]
        for i, code in enumerate(codes, start=3):
            ws3.cell(row=1, column=i, value=code.replace("total_", "Bo'lim "))
        for r, entry in enumerate(report["employee_breakdown"], start=2):
            ws3.cell(row=r, column=1, value=entry["full_name"])
            ws3.cell(row=r, column=2, value=entry["department"])
            for i, code in enumerate(codes, start=3):
                ws3.cell(row=r, column=i, value=entry["report"].get(code, 0))
        ws3.column_dimensions["A"].width = 30
        ws3.column_dimensions["B"].width = 25

    return wb


def workbook_to_bytes(wb) -> bytes:
    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
